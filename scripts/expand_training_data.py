"""Expand training data Excel with more paraphrases per intent."""
import openpyxl
import shutil

SRC = 'service_model_training_data_template.xlsx'
DST = 'service_model_training_data_template_expanded.xlsx'

# Map to ORIGINAL tag names from the Excel file
NEW_ROWS = {
    'password': {
        'type': 'answer',
        'pairs': [
            ("I can't remember my password", "パスワードが思い出せない"),
            ("How to change my login password?", "ログインパスワードの変更方法は？"),
            ("Reset password not working", "パスワードリセットがうまくいかない"),
            ("I need to update my password", "パスワードを更新したい"),
            ("Where do I change my password?", "パスワードはどこで変更できますか？"),
            ("My password expired", "パスワードの有効期限が切れました"),
            ("Account locked after wrong password", "パスワードを間違えてロックされた"),
            ("I keep getting password error", "パスワードエラーが出続けます"),
            ("Help me get back into my account", "アカウントに戻れるようにして"),
            ("Change login info", "ログイン情報を変更したい"),
            ("Forgot my credentials", "認証情報を忘れました"),
            ("How to set a new password", "新しいパスワードを設定するには"),
        ],
    },
    'subscription': {
        'type': 'answer',
        'pairs': [
            ("I want to stop paying for the service", "サービスの支払いを止めたい"),
            ("How to end my membership?", "会員を解約するには？"),
            ("Cancel my account subscription", "アカウントのサブスクをキャンセルしたい"),
            ("I don't want to renew", "更新したくありません"),
            ("Stop my monthly payment", "月額課金を止めてください"),
            ("Where can I cancel my plan?", "プランのキャンセルはどこで？"),
            ("I want to downgrade to free", "無料プランに戻したい"),
            ("End my premium subscription", "プレミアムの登録を終了したい"),
            ("Remove my subscription", "サブスクリプションを解除したい"),
            ("I no longer need this service", "このサービスはもう不要です"),
            ("Unsubscribe me from everything", "すべての登録を解除して"),
            ("Cancel renewal please", "更新をキャンセルしてください"),
        ],
    },
    'billing': {
        'type': 'answer',
        'pairs': [
            ("I see an unexpected charge", "予期しない請求があります"),
            ("Where can I view my invoices?", "請求書はどこで確認できますか？"),
            ("I was overcharged", "過剰請求されました"),
            ("When is my next billing date?", "次の請求日はいつですか？"),
            ("Can I get a refund?", "返金してもらえますか？"),
            ("How to switch to annual billing?", "年間払いに変更するには？"),
            ("My credit card was declined", "クレジットカードが拒否されました"),
            ("How much does the premium plan cost?", "プレミアムプランはいくらですか？"),
            ("I want to change my billing address", "請求先住所を変更したい"),
            ("Payment failed but money was taken", "支払い失敗なのにお金が引かれた"),
            ("How to apply a promo code?", "プロモーションコードの適用方法は？"),
            ("I need an itemized receipt", "明細付きの領収書が必要です"),
        ],
    },
    'howto': {
        'type': 'answer',
        'pairs': [
            ("How do I enable notifications?", "通知を有効にするには？"),
            ("Where are the settings?", "設定はどこですか？"),
            ("How to export my data?", "データをエクスポートするには？"),
            ("Can I use offline mode?", "オフラインモードは使えますか？"),
            ("How to change language?", "言語を変更するには？"),
            ("How to share my profile?", "プロフィールを共有するには？"),
            ("Is there a search function?", "検索機能はありますか？"),
            ("How do I delete my history?", "履歴を削除するには？"),
            ("Can I sync across devices?", "デバイス間で同期できますか？"),
            ("How to set up two-factor auth?", "二段階認証の設定方法は？"),
            ("How to customize my dashboard?", "ダッシュボードをカスタマイズするには？"),
            ("Can I change my username?", "ユーザー名を変更できますか？"),
        ],
    },
    'bug': {
        'type': 'support',
        'pairs': [
            ("Screen goes blank in settings", "設定を開くと画面が真っ白になる"),
            ("App freezes on loading screen", "ロード画面でフリーズする"),
            ("I keep getting error messages", "エラーメッセージが出続けます"),
            ("App crashes every time I log in", "ログインするたびにアプリが落ちる"),
            ("Buttons are not responding", "ボタンが反応しない"),
            ("Page doesn't load properly", "ページが正しく表示されない"),
            ("Lost my data after the update", "アップデート後にデータが消えた"),
            ("There's a display glitch", "表示にバグがあります"),
            ("Notifications stopped working", "通知が届かなくなった"),
            ("Search returns wrong results", "検索結果がおかしい"),
            ("Video won't play in the app", "アプリで動画が再生されない"),
            ("Error when I try to save", "保存しようとするとエラーが出る"),
        ],
    },
    'support': {
        'type': 'support',
        'pairs': [
            ("Let me speak to a manager", "マネージャーと話させてください"),
            ("I want to file a complaint", "苦情を申し立てたい"),
            ("This bot can't help me", "このボットでは解決できません"),
            ("Transfer me to a human agent", "人間のオペレーターに転送して"),
            ("I need to talk to someone real", "本物の人と話す必要があります"),
            ("Can I call customer service?", "カスタマーサービスに電話できますか？"),
            ("I want a phone call back", "折り返し電話がほしい"),
            ("Please escalate my issue", "問題をエスカレートしてください"),
            ("Is there a live chat option?", "ライブチャットはありますか？"),
            ("My problem is too complicated", "ボットでは対応できない問題です"),
            ("Get me a real support agent", "実際のスタッフにつないでください"),
            ("I need personalized help", "個別のサポートが必要です"),
        ],
    },
    'unknown': {
        'type': 'reject',
        'pairs': [
            ("What's the capital of France?", "フランスの首都はどこ？"),
            ("Play some music", "音楽を再生して"),
            ("How's the stock market today?", "今日の株式市場はどう？"),
            ("Write me an essay", "エッセイを書いて"),
            ("Translate this to Spanish", "これをスペイン語に翻訳して"),
            ("Book me a flight", "飛行機を予約して"),
            ("What's 2+2?", "2+2は？"),
            ("Set an alarm for 7am", "午前7時にアラームをセットして"),
            ("Who is the president?", "大統領は誰？"),
            ("Find me a restaurant nearby", "近くのレストランを探して"),
        ],
    },
}

def main():
    shutil.copy2(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb.active
    last_row = ws.max_row
    print(f"Original: {last_row - 2} data rows")

    added = 0
    for tag, data in NEW_ROWS.items():
        for en, ja in data['pairs']:
            last_row += 1
            ws.cell(row=last_row, column=1, value=tag)
            ws.cell(row=last_row, column=2, value=data['type'])
            ws.cell(row=last_row, column=3, value=en)
            ws.cell(row=last_row, column=4, value=ja)
            added += 1

    wb.save(DST)
    print(f"Added: {added} new rows ({added*2} samples EN+JA)")
    print(f"Total: {last_row - 2} data rows")
    print(f"Saved: {DST}")

if __name__ == '__main__':
    main()
